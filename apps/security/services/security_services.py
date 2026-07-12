# apps/security/services/security_services.py
import json
import logging
import sys
import traceback
from typing import List, Dict, Any, Optional, Tuple

# Librerías de Terceros
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Infraestructura Django Core
from django.contrib.auth import get_user_model
from django.db.models import Q
from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone


# Modelos y Componentes del Negocio Axentra OS
from apps.security.models import UserAppRole, AppModule
from apps.security.models.audit import SecurityAuditLog
from apps.security.services.permission_loader import get_app_permissions
from apps.security.utils.forensic_auditor import ForensicAuditor
from apps.security.utils.hierarchy_enforcer import HierarchyEnforcer

User = get_user_model()
logger = logging.getLogger(__name__)


class PermissionService:
    """Lógica transaccional centralizada para la inyección, mutación y auditoría de privilegios."""

    @staticmethod
    def authorize_new_user_entry(
        request,
        app_module: AppModule,
        user_id: str,
        rol_a_inyectar: str = "viewer",
    ) -> bool:
        """
        Incorpora un funcionario al padrón de una aplicación bajo filosofía Zero Trust.

        Reglas:
        - El rol se valida contra el ROLE_MAPPING específico de la app.
        - Permite roles funcionales por app: director_rh, oficial_rh, editor, reviewer, viewer, etc.
        - Sólo manager/root puede inyectar rol owner.
        - Owner de app puede inyectar roles funcionales no-owner.
        - Si el usuario ya tiene membresía activa en esa app, no se reinyecta.
        - Si existe membresía inactiva, la reactiva.
        """

        try:
            target_user = User.objects.get(
                id=user_id,
                is_deleted=False,
            )

            if target_user.is_manager or target_user.is_superuser:
                return False

            rol_limpio = str(rol_a_inyectar or "viewer").lower().strip()

            config_app = get_app_permissions(app_module.slug)

            permisos_disponibles = config_app.get("permissions", {}) or {}
            roles_app = config_app.get("roles", {}) or {}
            pesos_app = config_app.get("weights", {}) or {}

            if not roles_app:
                roles_app = {
                    "owner": list(permisos_disponibles.keys()),
                    "viewer": ["has_access_module"],
                }

            roles_validos = set(roles_app.keys())

            if rol_limpio not in roles_validos:
                rol_limpio = "viewer"

            is_manager_bypass = (
                getattr(request, "axentra_is_root", False)
                or getattr(request.user, "is_manager", False)
                or (
                    hasattr(request.user, "axentra_profile")
                    and getattr(request.user.axentra_profile, "is_root_admin", False)
                )
            )

            if rol_limpio == "owner" and not is_manager_bypass:
                return False

            if not is_manager_bypass:
                operador_es_owner_de_app = UserAppRole.objects.filter(
                    user=request.user,
                    app=app_module,
                    role="owner",
                    is_active=True,
                    is_deleted=False,
                ).exists()

                if not operador_es_owner_de_app:
                    return False

            with transaction.atomic():
                rol_existente = (
                    UserAppRole.objects
                    .filter(
                        user=target_user,
                        app=app_module,
                        is_deleted=False,
                    )
                    .first()
                )

                if rol_existente and rol_existente.is_active:
                    return False

                if rol_limpio == "owner":
                    llaves_finales = list(permisos_disponibles.keys())
                else:
                    llaves_finales = list(
                        roles_app.get(
                            rol_limpio,
                            ["has_access_module"],
                        )
                    )

                    if "has_access_module" not in llaves_finales:
                        llaves_finales.append("has_access_module")

                UserAppRole.objects.update_or_create(
                    user=target_user,
                    app=app_module,
                    defaults={
                        "role": rol_limpio,
                        "permissions_list": llaves_finales,
                        "is_active": True,
                        "is_deleted": False,
                    },
                )

                ForensicAuditor.registrar_evento(
                    request=request,
                    action_type=SecurityAuditLog.ActionTypes.CREATE,
                    module_component="MATRIZ_PERMISOS",
                    action_name="INYECCION_PERIMETRAL_FUNCIONARIO",
                    target_scope=(
                        f"Siembra inicial del funcionario {target_user.email} "
                        f"en {app_module.name} con rol [{rol_limpio.upper()}]."
                    ),
                    level=SecurityAuditLog.Levels.SUCCESS,
                    target_user=target_user,
                    search_target=str(target_user.id),
                    payload={
                        "initial_role": rol_limpio,
                        "role_weight": pesos_app.get(rol_limpio, 0),
                        "app_id": str(app_module.id),
                        "app_slug": app_module.slug,
                        "operador_id": str(request.user.id),
                        "operador_email": request.user.email,
                        "is_manager_bypass": is_manager_bypass,
                        "permissions_list": llaves_finales,
                    },
                )

            logger.info(
                f"🟢 CIBERSEGURIDAD: Funcionario {target_user.email} "
                f"sembrado en [{app_module.slug.upper()}] con rol [{rol_limpio.upper()}]."
            )

            return True

        except Exception as e:
            logger.error(
                f"❌ FALLO (authorize_new_user_entry): {str(e)}"
            )
            return False
    
    
    @staticmethod
    def save_matrix_permissions(request, target_user: Any, app_module: AppModule, nuevo_rol: str, llaves_encendidas: List[str], is_manager_bypass: bool = False) -> Tuple[bool, str]:
        """
        🚀 REFACTOR MASTER CENTRALIZADO BLINDADO: Sanea, valida jerarquía, calcula deltas y guarda.
        🛡️ ENFOQUE ZERO-TRUST ESTRICTO: Solo 'owner' hereda todos los permisos. ADMIN nace en ceros.
        🛰️ COMPILACIÓN FORENSE AUTÓNOMA: El servicio calcula el delta de forma nativa para blindar el payload_json.
        """
        try:
            rol_limpio = str(nuevo_rol).lower().strip()
            
            # 🪐 SNAPSHOT ANTES: Capturamos el estado actual real de la BD antes de alterarlo
            rol_actual_obj = UserAppRole.objects.filter(user=target_user, app=app_module).first()
            rol_anterior = rol_actual_obj.role if rol_actual_obj else 'ninguno'
            permisos_anteriores = list(rol_actual_obj.permissions_list or []) if rol_actual_obj else []

            # 🛡️ VALIDACIÓN DE ESCALAFÓN DE JERARQUÍA
            config_app = get_app_permissions(app_module.slug)
            weights_map = config_app.get('weights', {})
            permissions_pool = config_app.get('permissions', {})

            tiene_autoridad = HierarchyEnforcer.validar_autoridad_operador(
                request=request, target_user=target_user, app_module=app_module,
                nuevo_rol_slug=rol_limpio, weights_map=weights_map
            )
            if not tiene_autoridad:
                return False, "🚫 Violación de Escalafón: Tus privilegios locales no tienen el peso jerárquico requerido."

            # Saneamiento y filtrado de llaves inyectadas desde el POST contra el manifiesto
            lista_final_json = list(set([str(llave).strip() for llave in llaves_encendidas if llave]))
            if permissions_pool:
                lista_final_json = [l for l in lista_final_json if l in permissions_pool.keys()]

            # Regla de control Zero-Trust: Solo Owner clona la piscina completa
            if rol_limpio == 'owner' and permissions_pool:
                lista_final_json = list(permissions_pool.keys())
            
            # El token mínimo de entrada se amarra por diseño defensivo
            if 'has_access_module' not in lista_final_json:
                lista_final_json.append('has_access_module')

            # 🪐 ANÁLISIS FORENSE DELTA: Calculamos tokens ganados y perdidos de manera autónoma
            payload_delta = {
                'antes': {'role': rol_anterior, 'permissions': permisos_anteriores},
                'despues': {'role': rol_limpio, 'permissions': lista_final_json},
                'delta_cambios': {
                    'tokens_ganados': [p for p in lista_final_json if p not in permisos_anteriores],
                    'tokens_perdidos': [p for p in permisos_anteriores if p not in lista_final_json],
                    'rol_mutado': rol_anterior != rol_limpio
                }
            }

            # Persistencia física atómica en PostgreSQL
            with transaction.atomic():
                UserAppRole.objects.update_or_create(
                    user=target_user, app=app_module,
                    defaults={'role': rol_limpio, 'permissions_list': lista_final_json, 'is_active': True}
                )

            # 🪐 INYECTOR FORENSE CENTRALIZADO CON EL PAYLOAD COMPLETO
            action_code = "MUTACION_MATRIZ_BYPASS" if is_manager_bypass else "MUTACION_MATRIZ_ESTANDAR"
            level_status = SecurityAuditLog.Levels.SUCCESS if is_manager_bypass else SecurityAuditLog.Levels.INFO
            
            ForensicAuditor.registrar_evento(
                request=request,
                action_type=SecurityAuditLog.ActionTypes.ASSIGN,     # 🔤 Verbo Normalizado
                module_component="MATRIZ_PERMISOS",                  # 🗺️ Componente Local
                action_name=action_code,
                target_scope=f"Reconfiguración granular sobre la grilla de {target_user.email} en el módulo de {app_module.name}",
                level=level_status,
                target_user=target_user,
                search_target=target_user.id,
                payload=payload_delta                                # ◄── Guardado impecable en Postgres
            )

            msg = f"🔒 [NIVEL MAESTRO]: Reconfiguración por decreto completada." if is_manager_bypass else f"🔒 Matriz actualizada con éxito."
            return True, msg

        except Exception as e:
            logger.error(f"❌ FALLO TRANSACCIONAL (save_matrix_permissions): {str(e)}\n{traceback.format_exc()}")
            return False, f"Fallo interno de consistencia: {str(e)}"
        
        
    @staticmethod
    def exportar_auditoria_excel(request, filtros: dict) -> HttpResponse:
        """
        Exporta auditoría forense a Excel (.xlsx).

        Regla:
        - Si hay fecha_inicio o fecha_fin, respeta el rango.
        - Si no hay fechas, exporta las últimas 24 horas.
        - Incluye payload JSON como evidencia técnica.
        - Registra auditoría de la propia exportación.
        """

        try:
            import json
            from datetime import timedelta

            filtros = filtros or {}

            query = Q()

            fecha_inicio = filtros.get("fecha_inicio")
            fecha_fin = filtros.get("fecha_fin")

            if fecha_inicio or fecha_fin:
                if fecha_inicio:
                    query &= Q(created_at__date__gte=fecha_inicio)

                if fecha_fin:
                    query &= Q(created_at__date__lte=fecha_fin)

            else:
                hace_24_horas = timezone.now() - timedelta(hours=24)
                query &= Q(created_at__gte=hace_24_horas)

            app_namespace = filtros.get("app_namespace")
            action_type = filtros.get("action_type")
            level_status = filtros.get("level_status")
            search_target = filtros.get("search_target")
            operador = filtros.get("operador")

            if app_namespace:
                query &= Q(app_namespace=str(app_namespace).strip().lower())

            if action_type:
                query &= Q(action_type=str(action_type).strip().upper())

            if level_status:
                query &= Q(level_status=str(level_status).strip().upper())

            if search_target:
                query &= Q(search_target__icontains=str(search_target).strip())

            if operador:
                query &= Q(operator_user__email__icontains=str(operador).strip().lower())

            logs = (
                SecurityAuditLog.objects
                .filter(query)
                .select_related(
                    "operator_user",
                    "target_user",
                )
                .order_by("-created_at")
            )

            total_registros = logs.count()

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "AXENTRA FORENSIC AUDIT"
            ws.views.sheetView[0].showGridLines = True

            font_titulo = Font(
                name="Arial",
                size=11,
                bold=True,
                color="FFFFFF",
            )

            fill_titulo = PatternFill(
                start_color="0F172A",
                end_color="0F172A",
                fill_type="solid",
            )

            align_centro = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=False,
            )

            align_wrap = Alignment(
                vertical="top",
                wrap_text=True,
            )

            headers = [
                "IDPaquete (UUID)",
                "Fecha / Hora",
                "Criticidad",
                "Ecosistema / App",
                "Verbo (Action)",
                "Componente / Vista",
                "Operador (Email)",
                "Objetivo (Email)",
                "Descripción de Acción",
                "Criterio de Negocio (Target)",
                "Firma de Red (IP)",
                "Dispositivo / User-Agent",
                "Evidencia Técnica (JSON Payload)",
            ]

            ws.append(headers)

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = font_titulo
                cell.fill = fill_titulo
                cell.alignment = align_centro

            for log in logs:
                payload_raw = "{}"

                if log.payload_json:
                    if isinstance(log.payload_json, dict):
                        payload_raw = json.dumps(
                            log.payload_json,
                            ensure_ascii=False,
                            indent=2,
                        )
                    else:
                        payload_raw = str(log.payload_json)

                row = [
                    str(log.id),
                    timezone.localtime(log.created_at).strftime("%Y-%m-%d %H:%M:%S"),
                    str(log.level_status or "--"),
                    str(log.app_namespace or "core").upper(),
                    str(log.action_type or "--"),
                    str(log.module_component or "--"),
                    log.operator_user.email if log.operator_user else "SISTEMA",
                    log.target_user.email if log.target_user else "GLOBAL / SISTEMA",
                    str(log.action_name or "--"),
                    str(log.search_target or "--"),
                    str(log.ip_address or "--"),
                    str(log.user_agent or "--"),
                    payload_raw,
                ]

                ws.append(row)

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = align_wrap

            for col in ws.columns:
                max_len = max(
                    len(str(cell.value or ""))
                    for cell in col
                )

                col_letter = get_column_letter(col[0].column)
                header_value = col[0].value

                if header_value == "Evidencia Técnica (JSON Payload)":
                    ws.column_dimensions[col_letter].width = 55

                elif header_value == "Dispositivo / User-Agent":
                    ws.column_dimensions[col_letter].width = 55

                elif header_value == "Descripción de Acción":
                    ws.column_dimensions[col_letter].width = 42

                elif header_value == "Criterio de Negocio (Target)":
                    ws.column_dimensions[col_letter].width = 36

                else:
                    ws.column_dimensions[col_letter].width = min(
                        max(max_len + 3, 12),
                        35,
                    )

            ForensicAuditor.registrar_evento(
                request=request,
                action_type=SecurityAuditLog.ActionTypes.QUERY,
                module_component="AUDITORIA_FORENSE",
                action_name="EXPORTACION_EVIDENCIA_EXCEL",
                target_scope=(
                    "Exportación de evidencia forense de auditoría "
                    f"con {total_registros} registros."
                ),
                level=SecurityAuditLog.Levels.INFO,
                search_target="AUDIT_EXCEL_EXPORT",
                payload={
                    "total_registros": total_registros,
                    "filtros": filtros,
                    "operador_id": str(request.user.id),
                    "operador_email": request.user.email,
                },
            )

            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            response["Content-Disposition"] = (
                'attachment; filename="'
                f'axentra_audit_export_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
                '"'
            )

            wb.save(response)

            return response

        except Exception as e:
            logger.error(
                f"❌ FALLO AL EXPORTAR AUDITORÍA A EXCEL: {str(e)}",
                exc_info=True,
            )

            return HttpResponse(
                "Error interno al compilar el paquete de evidencia Excel.",
                status=500,
            )